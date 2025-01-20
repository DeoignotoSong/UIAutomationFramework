from openpyxl import load_workbook
from Core.Global import Data
from Utils.Log import log


class ExcelHandler:
    def __init__(self, project, scenario, default_devices):
        self._project = project
        self._scenario = scenario
        self._default_devices = default_devices

    def scenario_data(self):
        try:
            wb = load_workbook('./Projects/' + self._project + '/TestData/' + self._scenario + '.xlsx')
            Data.TEST_DATA[self._scenario] = {}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                Data.TEST_DATA[self._scenario][sheet_name] = {'In': {}, 'Out': {}}

                for row in ws.iter_rows():
                    if row[0].value is not None and row[0].value != '':
                        Data.TEST_DATA[self._scenario][sheet_name]['In'][row[0].value] = str(row[1].value)

            wb.close()
        except Exception as test_data_error:
            log('缺失{}场景测试数据Excel\n'.format(self._scenario) + str(test_data_error))
