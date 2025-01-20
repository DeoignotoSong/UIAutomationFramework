import re
import datetime

from openpyxl import Workbook, styles

from Core.Enum import TestStatus
from Core.Mail import Mail
from matplotlib import pyplot
from Core.Global import TestResult
from Core.Global import NameMap
from Core.ConfigOperation import ConfigOperation
from HelpTools.UIStatisticTools.CompletionSchedule import CompletionSchedule


class TestReport:
    def __init__(self, project_name, env='LAN', need_mail=False, is_show=False, project_during_time=0, **kwargs):

        print(TestResult.TEST_RESULTS)
        self.project_name = project_name
        self.need_mail = need_mail
        self.is_show = is_show
        self.env = env

        # 执行时间
        self.project_during_time = int(project_during_time)

        # 自定义报告参数
        config = ConfigOperation().load_config('./Projects/{project}/Base/config.ini'.format(project=self.project_name))

        self.svn_path = config.get('TestReport', 'svn_path')

        self.mail_title = config.get('TestReport', 'mail_title')
        self.html_title = config.get('TestReport', 'html_title')
        self.html_report_title = config.get('TestReport', 'html_report_title')
        self.html_js = config.get('TestReport', 'html_js')

        self.wb = Workbook()

        # 结果数据初始化
        self.success_scenario_count = 0
        self.failed_scenario_count = 0
        self.test_scenarios_result_mail_content = ''
        self.test_result_mail_content = '成功'
        self.test_result_color = '3DF44F'

        for scenario_name in TestResult.TEST_RESULTS:
            if TestResult.TEST_RESULTS[scenario_name]['scenario_result'] == TestStatus.Pass:
                self.success_scenario_count += 1
            else:
                self.failed_scenario_count += 1

        self.test_result = [
            ['测试结果', '测试数量'],
            ['成功', self.success_scenario_count],
            ['失败', self.failed_scenario_count]
        ]

        # 报告相关路径
        self.report_view_path = './Projects/' + self.project_name + '/TestReport/ReportView.html'
        self.report_image_path = './Projects/' + self.project_name + '/TestReport/image.png'
        self.report_excel_path = './Projects/' + self.project_name + '/TestReport/Report.xlsx'
        self.report_mail_path = './Projects/' + self.project_name + '/TestReport/Report.html'

        # 解决image中文乱码
        pyplot.rcParams['font.sans-serif'] = ['SimHei']
        # 更新svn路径
        # self._svn_handler()

    def generate_report(self,test_name):
        self.wb.remove(self.wb['Sheet'])

        # 创建总览页
        self.wb.create_sheet('总览')
        ws = self.wb['总览']

        ws['A1'] = '测试场景'
        ws['B1'] = '测试结果'

        test_scenarios_count = 1

        for scenario_name in TestResult.TEST_RESULTS.keys():
            self._scenario_details_excel_handle(scenario_name)

            self._scenario_summary_excel_handle(
                scenario_name,
                test_scenarios_count
            )

            self._scenario_details_mail_content_handle(scenario_name)

            test_scenarios_count += 1

        self.wb.save(self.report_excel_path)
        self._report_image_handler()

        if self.is_show:
            self._is_show_handler()

        if self.need_mail:
            self._mail_handle(
                self.env,
                test_name
            )

    def _scenario_details_excel_handle(self, scenario_name):
        """
        Scenario details page handler
        Generate scenario details worksheet content
        :param scenario_name: current scenario name
        """
        self.wb.create_sheet(scenario_name)
        ws = self.wb[scenario_name]

        ws['A1'] = '测试结果'
        ws['B1'] = '测试用例'
        ws['C1'] = '测试用例描述'
        ws['D1'] = '测试用例详情'
        ws['E1'] = '返回总览'
        ws['E1'].hyperlink = '#总览!A1'

        test_case_count = 0

        for item in TestResult.TEST_RESULTS[scenario_name]['test_case_details']:
            # 获取唯一的测试用例名，去掉重复测试用例后的计数后缀
            unique_test_case_name = re.sub(r'.\d+', '', item['test_step_name'])

            if item['test_result'] == TestStatus.Pass:
                ws['A' + str(test_case_count + 2)] = 'Pass'
                ws['A' + str(test_case_count + 2)].fill = styles.PatternFill(fill_type='solid', fgColor='00DB00')
            elif item['test_result'] == TestStatus.Failed:
                ws['A' + str(test_case_count + 2)] = 'Failed'
                ws['A' + str(test_case_count + 2)].fill = styles.PatternFill(fill_type='solid', fgColor='FF0000')
            elif item['test_result'] == TestStatus.Error:
                ws['A' + str(test_case_count + 2)] = 'Error'
                ws['A' + str(test_case_count + 2)].fill = styles.PatternFill(fill_type='solid', fgColor='FF0000')

            ws['D' + str(test_case_count + 2)] = item['result_detail']
            ws['B' + str(test_case_count + 2)] = item['test_step_name']
            try:
                ws['C' + str(test_case_count + 2)] = NameMap.TEST_CASE_MAP[unique_test_case_name]
            except Exception as error:
                ws['C' + str(test_case_count + 2)] = ''

            test_case_count += 1

        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 110

    def _scenario_summary_excel_handle(self, scenario_name, test_scenarios_count):
        """
        Generate scenario summary worksheet content
        :param scenario_name: current scenario name
        :param test_scenarios_count: current scenario index
        """
        # 生成总览页
        ws = self.wb['总览']

        ws['A' + str(test_scenarios_count + 1)] = scenario_name
        if TestResult.TEST_RESULTS[scenario_name]['scenario_result'] == TestStatus.Pass:
            ws['B' + str(test_scenarios_count + 1)] = 'Pass'
            ws['B' + str(test_scenarios_count + 1)].fill = \
                styles.PatternFill(fill_type='solid', fgColor='00DB00')
        else:
            ws['B' + str(test_scenarios_count + 1)] = 'Failed'
            ws['B' + str(test_scenarios_count + 1)].fill = \
                styles.PatternFill(fill_type='solid', fgColor='FF0000')

        ws['A' + str(test_scenarios_count + 1)].hyperlink = '#' + scenario_name + '!A1'

    def _scenario_details_mail_content_handle(self, scenario_name):
        """
        Generate mail content of scenario result list part.
        :param scenario_name: current scenario_name
        """
        # 生成测试报告邮件
        self.test_scenarios_result_mail_content += '<tr>\r\n'

        # 添加场景名称列
        self.test_scenarios_result_mail_content += '<td>\r\n'
        self.test_scenarios_result_mail_content += scenario_name + '\r\n'
        self.test_scenarios_result_mail_content += '</td>\r\n'

        # 添加场景描述列
        self.test_scenarios_result_mail_content += '<td>\r\n'
        self.test_scenarios_result_mail_content += NameMap.TEST_SCENARIO_MAP[scenario_name] + '\r\n'
        self.test_scenarios_result_mail_content += '</td>\r\n'

        # 添加场景结果列
        self.test_scenarios_result_mail_content += '<td>\r\n'

        if TestResult.TEST_RESULTS[scenario_name]['scenario_result'] == TestStatus.Pass:
            scenario_result_mail_content = '成功'
            scenario_color_code = '3DF44F'
        else:
            scenario_result_mail_content = '失败'
            scenario_color_code = 'F44336'

            self.test_result_mail_content = '失败'
            self.test_result_color = 'F44336'

        self.test_scenarios_result_mail_content += '<span style="color:#{};">{}</span>\r\n'.format(scenario_color_code,
                                                                                                   scenario_result_mail_content)
        self.test_scenarios_result_mail_content += '</td>\r\n'

        # 添加场景执行次数
        self.test_scenarios_result_mail_content += '<td>\r\n'
        self.test_scenarios_result_mail_content += '<span>{}</span>'.format(TestResult.TEST_RESULTS[scenario_name]['retry_time'])
        self.test_scenarios_result_mail_content += '</td>\r\n'

        # 添加场景执行时间
        self.test_scenarios_result_mail_content += '<td>\r\n'
        self.test_scenarios_result_mail_content += '<span>{} s</span>'.format(TestResult.TEST_RESULTS[scenario_name]['scenario_during_time'])
        self.test_scenarios_result_mail_content += '</td>\r\n'

    def _mail_handle(self, env, test_name):
        """
        Generate mail title and mail content, then send mail.
        """
        # 生成项目统计EXCEL
        test_scenarios_dir = './Projects/{project_name}/TestScenario'.format(project_name=self.project_name)
        test_cases_dir = './Projects/{project_name}/TestCases'.format(project_name=self.project_name)
        pages_dir = './Projects/{project_name}/Pages'.format(project_name=self.project_name)
        schedule_path = './HelpTools/UIStatisticTools/UITestFrameworkSchedule.xlsx'
        # cs = CompletionSchedule(schedule_path)
        # cs.main(test_scenarios_dir, test_cases_dir, pages_dir)
        # 初始化邮箱
        mail = Mail()

        with open('./Core/templates/MailContent.html', 'r', encoding='utf8') as mail_content_file:
            mail_content = ''

            for line in mail_content_file.readlines():
                mail_content += line + '\r\n'

        project_during_hour, project_during_minute, project_during_second = self.timestamp_transform(self.project_during_time)

        mail_content = mail_content.format(success_scenario_count=self.success_scenario_count,
                                           failed_scenario_count=self.failed_scenario_count,
                                           svn_path=self.svn_path,
                                           test_scenarios_result_mail_content=self.test_scenarios_result_mail_content,
                                           project_during_time_hour=project_during_hour,
                                           project_during_time_minute=project_during_minute,
                                           project_during_time_second=project_during_second)

        mail_title = self.mail_title + ' ' + test_name + '_' + self.test_result_mail_content

        with open(self.report_mail_path, 'w', encoding='utf8') as report_file:
            report_file.writelines(mail_content)

        mail.send_mail(mail_title,
                       mail_content,
                       self.report_image_path,
                       [self.report_excel_path, schedule_path])

    def _is_show_handler(self):
        """
        Generate is show mail title and mail content, then send mail.
        """
        with open('./Core/templates/PageViewTemplate.html', 'r', encoding='utf8') as mail_content_file:
            mail_content = ''

            for line in mail_content_file.readlines():
                mail_content += line + '\r\n'

        mail_content = mail_content.format(
            self.html_js,
            self.html_title,
            self.html_report_title,
            self.success_scenario_count,
            self.failed_scenario_count,
            self.svn_path,
            self.test_scenarios_result_mail_content
        )

        with open(self.report_view_path, 'w', encoding='utf8') as report_file:
            report_file.writelines(mail_content)

    def _report_image_handler(self):
        """
        Generate report image
        """
        # 调节图形大小
        pyplot.figure(figsize=(4, 4))
        # 定义标签
        labels = [u'Pass', u'Failed']
        # 区块数据
        sizes = [self.success_scenario_count,
                 self.failed_scenario_count]
        # 区块颜色
        colors = ['yellowgreen',
                  'red']
        # 将某一块分割出来，值越大分割出的间隙越大
        if self.success_scenario_count >= self.failed_scenario_count:
            explode = (0.02, 0)
        else:
            explode = (0, 0.02)

        image, text = pyplot.pie(
            sizes,
            explode=explode,
            labels=labels,
            colors=colors,
            labeldistance=1.2,
            shadow=False,
            startangle=90,
            pctdistance=0.6
        )

        # log(str(image))
        # log(str(text))

        pyplot.axis('equal')
        pyplot.legend()
        pyplot.savefig(self.report_image_path)

    def _svn_handler(self):
        month_folder = datetime.datetime.now().strftime('%Y.%m')
        day_folder = datetime.datetime.now().strftime('%m%d')

        self.svn_path = self.svn_path.format(month_folder,
                                             day_folder)

    @staticmethod
    def timestamp_transform(timestamp) -> tuple:
        hour = int(timestamp // 3600)
        minute = int(timestamp // 60 % 60)
        second = int(timestamp % 60)

        return str(hour), str(minute), str(second)
