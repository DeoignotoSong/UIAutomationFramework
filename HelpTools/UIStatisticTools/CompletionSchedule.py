import os
import sys
import re
from openpyxl import Workbook, utils
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

from openpyxl.drawing import image
from openpyxl.chart import BarChart, Series, Reference, BarChart3D
from openpyxl.chart.axis import DateAxis
from openpyxl import Workbook


class CompletionSchedule:
    """
    PC-UI自动化测试统计工具
    """
    def __init__(self, schedule_path):

        self.schedule_path = schedule_path
        self.test_scenarios = {}
        self.test_cases = {}
        self.pages = {}
        self.schedule_wb = Workbook()

    def find_py_files(self, file_dir, py_files_path=None):
        """
        遍历所有.py文件
        :param file_dir: 目录文件夹
        :param py_files_path: 用于递归存储.py文件路径
        :return: 文件路径集合: list
        """
        if py_files_path is None:
            py_files_path = []
        file_list = os.listdir(file_dir)

        for filename in file_list:
            file_path = os.path.join(file_dir, filename)

            if os.path.isdir(file_path):
                self.find_py_files(file_path, py_files_path=py_files_path)
            if file_path[-3:].upper() == '.PY':
                py_files_path.append(file_path)

        return py_files_path

    def get_scenarios(self, path):
        self.test_scenarios = self.test_scenarios.fromkeys(os.listdir(path))

    def get_scenario_info(self, file_path):
        """
        获取test_scenario信息
        :param file_path: 文件路径
        :return: tuple
            scenario_name：场景名称
            scenario_description: 场景描述
            test_cases: 包含用例
        """
        with open(file_path, encoding='utf-8') as scenario_file:
            file_content = scenario_file.read()

        # 获取场景名
        scenario_name = re.findall(r'(?<=def\s).*?(?=\()', file_content)
        if scenario_name:
            scenario_name = scenario_name[0]
        else:
            scenario_name = ''

        # 获取场景描述
        scenario_description = re.findall(r'(?<="""\n).*?(?=\n\s+:param\sargs:)', file_content)
        if scenario_description:
            scenario_description = scenario_description[0].strip()
        else:
            scenario_description = ''

        # 获取场景中用例名
        test_cases = re.findall(r'(?<==\s).*?(?=\(\*args,\s\*\*kwargs\))', file_content)

        self.test_scenarios[scenario_name] = {}
        self.test_scenarios[scenario_name]['description'] = scenario_description
        self.test_scenarios[scenario_name]['test_cases'] = test_cases

        return scenario_name, scenario_description, test_cases

    def get_cases_info(self, test_case_path):
        """
        获取test_case信息
        :param test_case_path: 文件路径
        :return: tuple
            case_name: 用例名称
            case_description: 用例描述
            case_pages: 包含页面
        """
        with open(test_case_path, encoding='utf-8') as case_file:
            file_content = case_file.read()

        # 获取用例名
        case_name = re.findall(r'(?<=def\s).*?(?=\()', file_content)
        if case_name:
            case_name = case_name[0]
        else:
            case_name = ''

        # 获取用例描述
        case_description = re.findall(r'(?<="""\n).*?(?=\n\s+:param)', file_content)
        if case_description:
            case_description = case_description[0].strip()
        else:
            case_description = ''

        # 获取包含页面
        case_pages = re.findall(r'(?<== ).*?(?=\(project\))', file_content)

        self.test_cases[case_name] = {}
        self.test_cases[case_name]['description'] = case_description
        self.test_cases[case_name]['pages'] = case_pages

        return case_name, case_description, case_pages

    def get_page_info(self, page_path):
        """
        获取Page信息
        :param page_path: 文件路径
        :return: tuple
            page_name: 页面名称
            page_description: 页面描述
            page_elements: 包含元素
        """
        if page_path.split('/')[-1].split('\\')[2][-3:].upper() == '.PY':
            item = page_path.split('/')[-1].split('\\')[1]
        else:
            item = page_path.split('/')[-1].split('\\')[1] + '-' + page_path.split('/')[-1].split('\\')[2]

        with open(page_path, encoding='utf-8') as page_file:
            file_content = page_file.read()

        # 获取页面名称
        page_name = re.findall(r'(?<=class ).*?(?=:)', file_content)
        if page_name:
            page_name = page_name[0].strip()
        else:
            page_name = ''

        # 获取页面描述
        page_description = re.findall(r'(?<="""\n).*?(?=\n)', file_content)
        if page_description:
            page_description = page_description[0].strip()
        else:
            page_description = re.findall(r"(?<='''\n).*?(?=\n)", file_content)
            if page_description:
                page_description = page_description[0].strip()
            else:
                page_description = ''

        # 获取页面地址
        page_url = re.findall(r'(?<={}\n).*?(?=\n"""\n)'.format(page_description), file_content)
        # page_url = [i for i in page_url if i != '']
        if page_url:
            page_url = page_url[0].strip()
            pass
        else:
            page_url = re.findall(r"(?<={}\n).*?(?=\n'''\n)".format(page_description), file_content)
            if page_url:
                page_url = page_description[0].strip()
            else:
                page_url = ''

        # 获取页面元素
        page_elements = re.findall(r'(?<=# ).*?(?=\n\s+self._)', file_content)

        if item not in self.pages.keys():
            self.pages[item] = {}

        self.pages[item][page_name] = {}
        self.pages[item][page_name]['description'] = page_description
        self.pages[item][page_name]['url'] = page_url
        self.pages[item][page_name]['elements'] = page_elements

        return page_name, page_description, page_elements

    def generate_overview(self):
        """
        生成总览sheet
        :return:
        """
        self.schedule_wb.create_sheet('总览')
        ws = self.schedule_wb['总览']

        pages_count = 0
        for item in self.pages.keys():
            pages_count += self.pages[item].__len__()

        rows = [
            ('分类', '完成数量'),
            ("测试场景", self.test_scenarios.__len__()),
            ("测试用例", self.test_cases.__len__()),
            ("测试页面", pages_count)
        ]
        for row in rows:
            ws.append(row)

        self.bar_3d(title='PC-UI计划统计表', data_max_col=2, data_max_row=4, title_max_row=4, location='A9')

    def generate_scenarios_sheet(self):
        """
        生成测试场景sheet
        :return:
        """
        self.schedule_wb.create_sheet('测试场景')
        ws = self.schedule_wb['测试场景']

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 80

        titles = ['测试场景', '场景描述', '用例数量', '包含用例']

        for index in range(0, titles.__len__()):
            ws.cell(row=1, column=index + 1, value=titles[index]).alignment = Alignment(horizontal='center', vertical='center')

        row = 2
        for scenario_name in self.test_scenarios.keys():
            test_cases = ',\n'.join(self.test_scenarios[scenario_name]['test_cases'])
            ws.cell(row=row, column=1, value=scenario_name)
            ws.cell(row=row, column=2, value=self.test_scenarios[scenario_name]['description'])
            ws.cell(row=row, column=3, value=self.test_scenarios[scenario_name]['test_cases'].__len__()).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=4, value=test_cases)
            row += 1

        ws.freeze_panes = 'A2'

    def generate_cases_sheet(self):
        """
        生成测试用例sheet
        :return:
        """
        self.schedule_wb.create_sheet('测试用例')
        ws = self.schedule_wb['测试用例']

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 60

        titles = ['测试用例', '用例描述', '页面数量', '包含页面']

        for index in range(0, titles.__len__()):
            ws.cell(row=1, column=index + 1, value=titles[index]).alignment = Alignment(horizontal='center', vertical='center')

        row = 2
        for case_name in self.test_cases.keys():
            pages = ',\n'.join(self.test_cases[case_name]['pages'])
            ws.cell(row=row, column=1, value=case_name)
            ws.cell(row=row, column=2, value=self.test_cases[case_name]['description'])
            ws.cell(row=row, column=3, value=self.test_cases[case_name]['pages'].__len__()).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=4, value=pages)
            row += 1

        ws.freeze_panes = 'A2'

    def generate_pages_sheet(self):
        """
        生成测试页面sheet
        :return:
        """
        self.schedule_wb.create_sheet('测试页面')
        ws = self.schedule_wb['测试页面']

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 80

        titles = ['页面分类', '测试页面', '页面描述', '页面URL', '元素数量', '包含元素']

        for index in range(0, titles.__len__()):
            ws.cell(row=1, column=index + 1, value=titles[index]).alignment = Alignment(horizontal='center', vertical='center')

        row = 2
        for item in self.pages.keys():
            page_count = 0
            for page in self.pages[item].keys():
                pages = ',\n'.join(self.pages[item][page]['elements'])
                ws.cell(row=row, column=2, value=page)
                ws.cell(row=row, column=3, value=self.pages[item][page]['description'])
                ws.cell(row=row, column=4, value=self.pages[item][page]['url'])
                ws.cell(row=row, column=5, value=self.pages[item][page]['elements'].__len__()).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=6, value=pages)
                row += 1
                page_count += 1

            ws.merge_cells(start_row=row-page_count, start_column=1, end_row=row-1, end_column=1)
            ws.cell(row=row-page_count, column=1, value=item).alignment = Alignment(horizontal='center', vertical='center')

        ws.freeze_panes = 'A2'

    def generate_schedule(self):
        """
        生成excel
        :return:
        """
        self.schedule_wb.remove_sheet(self.schedule_wb['Sheet'])
        self.generate_overview()
        self.generate_scenarios_sheet()
        self.generate_cases_sheet()
        self.generate_pages_sheet()
        self.schedule_wb.save(self.schedule_path)
        self.schedule_wb.close()

    def bar_3d(self, title, data_max_col, data_max_row, title_max_row, location):
        """
        生成3D柱状图
        数据默认从第一行、第二列开始
        标题默认从第一列、第二行开始
        :param title: 柱状图标题
        :param data_max_col: 数据最大列数
        :param data_max_row: 数据最大行数
        :param title_max_row: 标题最大行数
        :param location: 图标展示位置：A9
        :return:
        """
        ws = self.schedule_wb['总览']
        data = Reference(ws, min_col=2, min_row=1, max_col=data_max_col, max_row=data_max_row)
        titles = Reference(ws, min_col=1, min_row=2, max_row=title_max_row)
        chart = BarChart3D()
        chart.title = title
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        ws.add_chart(chart, location)

    def main(self, test_scenarios_dir, test_cases_dir, pages_dir):
        """
        执行入口
        :param test_scenarios_dir: 测试场景文件目录
        :param test_cases_dir: 测试用例文件目录
        :param pages_dir: 页面文件目录
        :return:
        """
        test_scenarios_path = self.find_py_files(test_scenarios_dir)
        test_cases_path = self.find_py_files(test_cases_dir)
        pages_path = self.find_py_files(pages_dir)
        print(test_scenarios_path)
        print(test_cases_path)
        print(pages_path)

        for test_scenario_path in test_scenarios_path:
            scenario_info = self.get_scenario_info(test_scenario_path)

        for test_case_path in test_cases_path:
            case_info = self.get_cases_info(test_case_path)

        for page_path in pages_path:
            page_info = self.get_page_info(page_path)

        self.generate_schedule()


if __name__ == '__main__':

    # 文件目录
    test_scenarios_dir = './Projects/PC/TestScenario'
    test_cases_dir = './Projects/PC/TestCases'
    pages_dir = './Projects/PC/Pages'
    schedule_path = './HelpTools/UIStatisticTools/UITestFrameworkSchedule.xlsx'

    cs = CompletionSchedule(schedule_path)
    cs.main(test_scenarios_dir, test_cases_dir, pages_dir)





